import openpyxl
import datetime

settings = openpyxl.open('avito-settings.xlsx', read_only=True).active
products = openpyxl.open('avito-products.xlsx', read_only=True).active
cities = openpyxl.open('avito-cities.xlsx', read_only=True).active

def multiplication():
    autoload_wb = openpyxl.open('avito-autoload.xlsx')
    autoload = autoload_wb.active
    autoload_row_max = 0
    setting = {}
    for row in range(3, settings.max_row + 1):
        setting_dict = {}
        for column in range(2, settings.max_column + 1):
            setting_dict[settings.cell(2, column).value] = settings.cell(row, column).value
        autoload_row_max += settings.cell(row, 4).value
        setting_dict['count_posting_index'] = 1
        setting[settings.cell(row, 1).value] = setting_dict
    for column in range(1, products.max_column + 1):
        autoload.cell(1, column=column).value = products.cell(1, column).value
    for column in range(1, cities.max_column + 1):
        autoload.cell(1, products.max_column + column).value = cities.cell(2, column).value
    autoload_row = 2
    for product_row in range(2, products.max_row):
        for city_row in range(3, cities.max_row):
            if cities.cell(city_row, 5).value == 1:
                
                if setting[products.cell(product_row, 3).value]['count_posting_index'] <= int(setting[products.cell(product_row, 3).value]['count_posting']):

                    for product_column in range(1, products.max_column + 1):
                        autoload.cell(autoload_row, product_column).value = products.cell(product_row, product_column).value
                    
                    for city_column in range(1, cities.max_column + 1):
                        autoload.cell(autoload_row, products.max_column + city_column).value = cities.cell(city_row, city_column).value
                    
                    setting[autoload.cell(autoload_row, 3).value]['count_posting_index'] += 1
                    
                    autoload.cell(autoload_row, 1).value = autoload_row - 1
                    autoload_row += 1

    autoload_wb.save('avito-autoload.xlsx')
    autoload_wb.close()

def add_id():
    autoload_wb = openpyxl.open('avito-autoload.xlsx')
    autoload = autoload_wb.active
    for row in range(2, autoload.max_row + 1):
        if len(str(autoload.cell(row, 1).value)) == 1:
            str_null = '0000'
        if len(str(autoload.cell(row, 1).value)) == 2:
            str_null = '000'
        if len(str(autoload.cell(row, 1).value)) == 3:
            str_null = '00'
        if len(str(autoload.cell(row, 1).value)) == 4:
            str_null = '0'
        autoload.cell(row, 5).value = str(autoload.cell(row, 4).value) + '-' + str_null + str(autoload.cell(row, 1).value) + '-' + str(autoload.cell(row, 3).value)
    autoload_wb.save('avito-autoload.xlsx')
    autoload_wb.close()

def add_city_in_title():
    pass

def add_title_in_text():
    pass

def add_address():
    autoload_wb = openpyxl.open('avito-autoload.xlsx')
    autoload = autoload_wb.active

    for row in range(2, autoload.max_row + 1):
        autoload.cell(row, 13).value = autoload.cell(row, 30).value

    autoload_wb.save('avito-autoload.xlsx')
    autoload_wb.close()

def add_images():
    autoload_wb = openpyxl.open('avito-autoload.xlsx')
    autoload = autoload_wb.active
    for row in range(2, autoload.max_row + 1):
        autoload.cell(row, 26).value = str(autoload.cell(row, 24).value) + str(autoload.cell(row, 5).value) + str(autoload.cell(row, 25).value)
    autoload_wb.save('avito-autoload.xlsx')
    autoload_wb.close()

def add_date():
    autoload_wb = openpyxl.open('avito-autoload.xlsx')
    autoload = autoload_wb.active
    autoload_row_max = 0
    setting = {}
    
    for row in range(3, settings.max_row + 1):
        setting_dict = {}
        for column in range(2, settings.max_column + 1):
            setting_dict[settings.cell(2, column).value] = settings.cell(row, column).value
        autoload_row_max += settings.cell(row, 4).value
        setting_dict['count_posting_index'] = 1
        setting[settings.cell(row, 1).value] = setting_dict
    autoload_index = 2
    for setting_row in range(3, settings.max_row + 1):
        category_name = settings.cell(setting_row, 1).value
        if autoload_index <= autoload_row_max:
            str_date = str(settings.cell(setting_row, 2).value).split(',')
            str_time = str(settings.cell(setting_row, 3).value).split(',')
            if str_date and str_time:
                datetime_index = datetime.datetime(
                    int(str_date[0]), 
                    int(str_date[1]), 
                    int(str_date[2]), 
                    int(str_time[0]), 
                    int(str_time[1]), 
                    int(str_time[2])
                )
            for count_day in range(1, setting[category_name]['count_days'] + 1):
                for count_hours_in_day in range (1, setting[category_name]['count_hours_in_day'] + 1):
                    for count_posting_in_hours in range(1, setting[category_name]['count_posting_in_hours'] + 1):
                        if autoload_index <= autoload_row_max + 1:
                            if setting[category_name]['count_posting_index'] <= setting[category_name]['count_posting']:
                                autoload.cell(autoload_index, 6).value = datetime_index.strftime('%Y-%m-%dT%H:%M:%S+03:00')
                                setting[category_name]['count_posting_index'] += 1
                                autoload_index += 1
                    datetime_index += datetime.timedelta(hours=1)
                datetime_index += datetime.timedelta(days=1)
                datetime_index = datetime.datetime(
                    year=datetime_index.year,
                    month=datetime_index.month,
                    day=datetime_index.day,
                    hour=int(str_time[0]), 
                    minute=int(str_time[1]), 
                    second=int(str_time[2])
                )
    autoload_wb.save('avito-autoload.xlsx')
    autoload_wb.close()
def main():
    print("[0] - Завершить программу")
    print("[1] - Перемножить Товары на Города")
    print("[2] - Добавить Id")
    print("[3] - Добавить Город в Заголовок")
    print("[4] - Добавить Заголовок в Описание")
    print("[5] - Добавить Адрес")
    print("[6] - Добавить Изображения")
    print("[7] - Добавить Даты")

    btn = input("Выберите действие над данными: ")

    match btn:
        case "0":
            exit
        case "1":
            multiplication()
            main()
        case "2":
            add_id()
            main()
        case "3":
            add_city_in_title()
            main()
        case "4":
            add_title_in_text()
            main()
        case "5":
            add_address()
            main()    
        case "6":
            add_images()
            main()
        case "7":
            add_date()
            main() 

#multiplication()
#add_id()
#add_address()
#add_images()
add_date()